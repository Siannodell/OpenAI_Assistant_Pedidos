import base64

import openai
import streamlit as st
from bs4 import BeautifulSoup
from faker.decode import unidecode
from streamlit.components.v1 import html
import requests
import pdfkit
import time
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from PIL import Image
import urllib
import pandas as pd
import threading

from packages.chatOpenAi import chat_openai

load_dotenv()
#id do assistente
assistant_id = "asst_RyDmETRf7S9E7gPmXje6HKwD"
def check_streamlit():
    thread = threading.current_thread()
    return type(thread).__module__.startswith('streamlit.')

# inicializa cliente openai
client = openai

# inicializa a sessão para ler os ids
if "file_id_list" not in st.session_state:
    st.session_state.file_id_list = []

if "start_chat" not in st.session_state:
    st.session_state.start_chat = False

if "thread_id" not in st.session_state:
    st.session_state.thread_id = None

# titulo e icone da página
# Função para converter XLSX pra PDF

def convert_xlsx_to_json(input_path, output_path) :
    read_file = pd.read_excel(input_path)

    read_file.to_json(output_path)

def convert_xlsx_to_pdf(input_path, output_path):
    workbook = load_workbook(input_path)
    sheets = workbook.sheetnames

    pdf = canvas.Canvas(output_path, pagesize=letter)

    for sheet_name in sheets:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                pdf.drawString(cell.column * 50, letter[1] - cell.row * 10, str(cell.value))

    pdf.save()


perguntas = [
    "Qual faixa etária apresentou o maior volume de pedidos e qual foi o valor médio destes pedidos?",
    "Há diferenças significativas nos padrões de compra entre os diferentes gêneros listados no documento?",
    "Qual foi o ticket médio dos pedidos aprovados comparado com os pedidos não aprovados? Isso pode indicar alguma tendência ou comportamento específico dos consumidores?",
    "Qual é a taxa de aprovação dos pedidos recebidos e como ela se distribui entre as diferentes cidades ou estados?",
    "A localização impacta o valor médio dos pedidos ou a preferência por formas de pagamento?",
    "Comparando os dados de agosto de 2023 com meses anteriores, existe alguma tendência de crescimento ou decréscimo nas transações?",
]

pergunta_ = ""




def download_file(file) :
    file = urllib.request.urlopen(file).read()
    return BytesIO(file)

# Função pra enviar arquivo convertido pra OpenAI
def upload_to_openai(filepath):
    with open(filepath, "rb") as file:
        response = openai.files.create(file=file.read(), purpose="assistants")
    return response.id

#local
api_key = os.getenv("OPENAI_API_KEY")
#git
#api_key = st.secrets.OpenAIAPI.openai_api_key
#Faça a apresentação do valor aprovado e valor recebido por mes dos ultimos 12 meses e utilize gráficos em visualizações para facilitar a interpretação dos resultados. Forneça insights sobre o que pode ter impactados os 3 meses com menor volume de vendas e o que pode tem impactado os 3 meses com maior volume de vendas. Forneça recomendações para aumentar as vendas aprovadas.


if api_key:
    openai.api_key = api_key

#st.sidebar.write("<a style='color:white'  href='https://tecnologia2.chleba.net/_ftp/chatgpt/BotasVentoPedidos.xlsx' id='baixarArquivo'>[Baixe o arquivo para fazer a análise]</a>", unsafe_allow_html=True)

#uploaded_file = st.sidebar.file_uploader("Envie um arquivo", key="file_uploader")
uploaded_file = download_file("https://tecnologia2.chleba.net/_ftp/chatgpt/BotasVentoPedidos.xlsx")
# Botão para iniciar o chat

st.sidebar.write('<style>p, ol, ul, dl {font-size:0.9rem}</style>', unsafe_allow_html=True)

if not st.session_state.start_chat:
    if True:
        #if uploaded_file:
        ds = client.beta.assistants.files.list(assistant_id=assistant_id)
        if ds:
            for file in ds:
                st.session_state.file_id_list.append(file.id)
            #st.sidebar.write(f"ID do arquivo: {additional_file_id}")

        # Mostra os ids
        if st.session_state.file_id_list:
            #st.sidebar.write("IDs dos arquivos enviados:")
            for file_id in st.session_state.file_id_list:
                #st.sidebar.write(file_id)
                # Associa os arquivos ao assistente
                assistant_file = client.beta.assistants.files.create(
                    assistant_id=assistant_id,
                    file_id=file_id
                )


        # Verifica se o arquivo foi upado antes de iniciar
        if st.session_state.file_id_list:
            st.session_state.start_chat = True
            # Cria a thread e guarda o id na sessão
            thread = client.beta.threads.create()
            st.session_state.thread_id = thread.id
            #st.write("id da thread: ", thread.id)
        else:
            st.sidebar.warning("Por favor, clique em \"Iniciar análise\" iniciar o chat")


if st.session_state.start_chat:
    on = st.sidebar.toggle('Ver sugestões de perguntas', value=True)
    search = st.sidebar.text_input("Pesquisar perguntas sugeridas")

    if on:
        for indice, pergunta in enumerate(perguntas):
            # st.sidebar.write(f"<a style=\"color:white;display:flex;align-items:center;gap:26px;text-decoration:none\" target=\"_self\" id=\"pergunta{indice}\" href=\"javascript:(function(){{var conteudo = document.getElementById('pergunta{indice}').innerText; navigator.clipboard.writeText(conteudo).then(function() {{ console.log('Conteúdo copiado para a área de transferência: ' + conteudo); }}, function(err) {{ console.error('Erro ao copiar conteúdo: ', err); }});}})()\">{pergunta}<span>{icon_copy}</span></a>", unsafe_allow_html=True)
            if unidecode(search.lower()) in unidecode(pergunta.lower()):
                if st.sidebar.button(f"{pergunta}"):
                    pergunta_ = pergunta


    st.sidebar.write('<style>label[data-baseweb="checkbox"] > div > div {background: #282828}</style>', unsafe_allow_html=True)
# Define a função para iniciar


# Interface do chat
st.subheader("ANÁLISE DE PEDIDOS")
#st.write("Este chat usa a API da OpenAI para gerar respostas.")
st.session_state.is_running = False
# Só vai mostrar o chat se for iniciado
if st.session_state.start_chat:
    chat_openai(pergunta_, client, assistant_id)

else:
    # Prompt pra iniciar o chat
    st.write("Por favor, selecione o(s) arquivo(s) e clique em *iniciar chat* para gerar respostas")
